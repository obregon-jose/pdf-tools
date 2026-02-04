import os
import re
import json
import threading
import time
import random
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook, Workbook
from pathlib import Path
import requests

# Archivo para guardar el token
TOKEN_FILE = Path.home() / ".horus_token.json"


class HorusApp(ctk.CTkFrame):
    """Aplicación para consulta de afiliados en Horus Health."""
    
    DOC_TYPES = {"CC": 1, "TI": 2, "RC": 3, "CE": 5, "CN": 12}
    LOGIN_URL = "https://backend.horus-health.com/api/auth/validar-usuario"
    BASE_URL = "https://backend.horus-health.com/api/afiliados/consultar-afiliado"
    ME_URL = "https://backend.horus-health.com/api/auth/me"
    
    def __init__(self, master, go_home=None):
        super().__init__(master)
        self.token = None
        self.session = requests.Session()
        self.session.headers.update({"Accept": "application/json", "Content-Type": "application/json"})
        self.excel_docs = []
        self.is_processing = False
        self._build_ui()
        # Verificar token automáticamente al abrir
        threading.Thread(target=self._auto_validate_token, daemon=True).start()
    
    def _load_token(self):
        """Carga el token guardado."""
        try:
            if TOKEN_FILE.exists():
                with open(TOKEN_FILE, "r") as f:
                    return json.load(f).get("token")
        except:
            pass
        return None
    
    def _save_token(self, token):
        """Guarda el token."""
        try:
            with open(TOKEN_FILE, "w") as f:
                json.dump({"token": token}, f)
        except:
            pass
    
    def _delete_token(self):
        """Elimina el token guardado."""
        try:
            if TOKEN_FILE.exists():
                TOKEN_FILE.unlink()
        except:
            pass
    
    def _auto_validate_token(self):
        """Valida automáticamente el token al iniciar la app."""
        saved_token = self._load_token()
        
        if not saved_token:
            self.after(0, lambda: self.progress_lbl.configure(text="⏳ Esperando inicio de sesión...", text_color="gray"))
            return
        
        self.after(0, lambda: self.progress_lbl.configure(text="🔄 Verificando sesión guardada...", text_color="orange"))
        
        try:
            res = requests.get(
                self.ME_URL,
                headers={"Authorization": f"Bearer {saved_token}"},
                timeout=10
            )
            
            if res.status_code == 200:
                self.token = saved_token
                self.session.headers.update({"Authorization": f"Bearer {self.token}"})
                self.after(0, lambda: self._set_connected(True))
                self.after(0, lambda: self.progress_lbl.configure(text="✅ Sesión restaurada automáticamente.", text_color="#27ae60"))
            else:
                self._delete_token()
                self.after(0, lambda: self._set_connected(False))
                self.after(0, lambda: self.progress_lbl.configure(text="⚠️ Sesión expirada. Inicia sesión nuevamente.", text_color="orange"))
        except:
            self.after(0, lambda: self.progress_lbl.configure(text="⚠️ Sin conexión. Verifica tu internet.", text_color="orange"))
    
    def _build_ui(self):
        """Construye toda la interfaz de usuario."""
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        
        self._build_login_panel()
        self._build_input_panel()
        self._build_main_panel()
        self._build_progress_panel()
    
    def _build_login_panel(self):
        """Panel de login."""
        panel = ctk.CTkFrame(self)
        panel.grid(row=0, column=0, sticky="ew", padx=6, pady=6)
        panel.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(panel, text="📧 Correo:", width=90, anchor="w", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=(6,4), pady=4, sticky="w")
        self.email_entry = ctk.CTkEntry(panel, placeholder_text="correo@ejemplo.com")
        self.email_entry.grid(row=0, column=1, padx=4, pady=4, sticky="ew")
        
        ctk.CTkLabel(panel, text="🔒 Contraseña:", width=90, anchor="w", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, padx=(6,4), pady=4, sticky="w")
        self.password_entry = ctk.CTkEntry(panel, placeholder_text="••••••••", show="•")
        self.password_entry.grid(row=1, column=1, padx=4, pady=4, sticky="ew")
        
        login_frame = ctk.CTkFrame(panel, fg_color="transparent")
        login_frame.grid(row=0, column=2, rowspan=2, padx=6, pady=4)
        
        self.login_btn = ctk.CTkButton(login_frame, text="🔌 Conectar", width=130, command=self._do_login, fg_color="#1f6feb")
        self.login_btn.pack(pady=(0,4))
        
        self.status_lbl = ctk.CTkLabel(login_frame, text="⚫ DESCONECTADO", text_color="#e74c3c", font=ctk.CTkFont(size=11, weight="bold"))
        self.status_lbl.pack()
    
    def _build_input_panel(self):
        """Panel de entrada y acciones."""
        panel = ctk.CTkFrame(self)
        panel.grid(row=1, column=0, sticky="ew", padx=6, pady=6)
        panel.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(panel, text="📂 Modo:", width=90, anchor="w", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=(6,4), pady=4, sticky="w")
        self.mode_selector = ctk.CTkSegmentedButton(panel, values=["Manual", "Excel"], command=self._on_mode_change, width=180)
        self.mode_selector.grid(row=0, column=1, padx=4, pady=4, sticky="w")
        self.mode_selector.set("Manual")
        
        self.excel_btn = ctk.CTkButton(panel, text="📁 Seleccionar Excel", width=140, command=self._select_excel)
        self.excel_lbl = ctk.CTkLabel(panel, text="", text_color="gray", anchor="w")
        
        btn_frame = ctk.CTkFrame(panel, fg_color="transparent")
        btn_frame.grid(row=1, column=0, columnspan=4, sticky="ew", padx=4, pady=(4,6))
        btn_frame.grid_columnconfigure((0,1,2), weight=1)
        
        self.query_btn = ctk.CTkButton(btn_frame, text="🔍 Consultar Afiliados", command=self._do_query, fg_color="#1f6feb")
        self.query_btn.grid(row=0, column=0, padx=4, pady=4, sticky="ew")
        
        self.export_btn = ctk.CTkButton(btn_frame, text="📥 Exportar Excel", command=self._do_export, fg_color="#27ae60")
        self.export_btn.grid(row=0, column=1, padx=4, pady=4, sticky="ew")
        
        self.clear_btn = ctk.CTkButton(btn_frame, text="🧹 Limpiar Todo", command=self._do_clear, fg_color="#e74c3c")
        self.clear_btn.grid(row=0, column=2, padx=4, pady=4, sticky="ew")
    
    def _build_main_panel(self):
        """Panel principal."""
        panel = ctk.CTkFrame(self)
        panel.grid(row=2, column=0, sticky="nsew", padx=6, pady=6)
        panel.grid_columnconfigure(0, weight=1)
        panel.grid_columnconfigure(1, weight=2)
        panel.grid_rowconfigure(0, weight=1)
        
        self._build_docs_panel(panel)
        self._build_results_panel(panel)
    
    def _build_docs_panel(self, parent):
        """Panel de documentos."""
        panel = ctk.CTkFrame(parent)
        panel.grid(row=0, column=0, sticky="nsew", padx=(0,3), pady=0)
        panel.grid_columnconfigure(0, weight=1)
        panel.grid_rowconfigure(1, weight=1)
        
        self.docs_lbl = ctk.CTkLabel(panel, text="📑 Documentos: 0", font=ctk.CTkFont(size=12, weight="bold"), anchor="w")
        self.docs_lbl.grid(row=0, column=0, sticky="w", padx=6, pady=(6,2))
        
        self.docs_text = ctk.CTkTextbox(panel, font=ctk.CTkFont(family="Consolas", size=12), border_width=2)
        self.docs_text.grid(row=1, column=0, sticky="nsew", padx=6, pady=(2,6))
        self.docs_text.bind("<KeyRelease>", self._update_docs_count)
    
    def _build_results_panel(self, parent):
        """Panel de resultados."""
        panel = ctk.CTkFrame(parent)
        panel.grid(row=0, column=1, sticky="nsew", padx=(3,0), pady=0)
        panel.grid_columnconfigure(0, weight=1)
        panel.grid_rowconfigure(1, weight=1)
        
        header = ctk.CTkFrame(panel, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=6, pady=(6,2))
        header.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(header, text="📊 Resultados", font=ctk.CTkFont(size=12, weight="bold"), anchor="w").grid(row=0, column=0, sticky="w")
        self.results_count_lbl = ctk.CTkLabel(header, text="0 registros", text_color="gray", anchor="e")
        self.results_count_lbl.grid(row=0, column=1, sticky="e")
        
        table_frame = ctk.CTkFrame(panel)
        table_frame.grid(row=1, column=0, sticky="nsew", padx=6, pady=(2,6))
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)
        
        self._setup_table_style()
        
        cols = ("documento", "nombre", "ips", "estado")
        self.table = ttk.Treeview(table_frame, columns=cols, show="headings", style="Dark.Treeview")
        
        self.table.heading("documento", text="Documento")
        self.table.heading("nombre", text="Nombre Completo")
        self.table.heading("ips", text="IPS")
        self.table.heading("estado", text="Estado")
        
        self.table.column("documento", width=100, minwidth=80)
        self.table.column("nombre", width=200, minwidth=150)
        self.table.column("ips", width=150, minwidth=100)
        self.table.column("estado", width=100, minwidth=80)
        
        scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.table.yview)
        scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.table.xview)
        self.table.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.table.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
    
    def _build_progress_panel(self):
        """Panel de progreso."""
        panel = ctk.CTkFrame(self)
        panel.grid(row=3, column=0, sticky="ew", padx=6, pady=6)
        
        self.progressbar = ctk.CTkProgressBar(panel, height=10, corner_radius=5)
        self.progressbar.pack(fill="x", padx=6, pady=(6,4))
        self.progressbar.set(0)
        
        self.progress_lbl = ctk.CTkLabel(panel, text="🔄 Iniciando...", text_color="orange", anchor="w")
        self.progress_lbl.pack(fill="x", padx=6, pady=(0,6))
    
    def _setup_table_style(self):
        """Estilo oscuro para tabla."""
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Dark.Treeview", background="#2b2b2b", foreground="white", fieldbackground="#2b2b2b", borderwidth=0, rowheight=26)
        style.configure("Dark.Treeview.Heading", background="#3b3b3b", foreground="white", borderwidth=1)
        style.map("Dark.Treeview", background=[("selected", "#1f6feb")], foreground=[("selected", "white")])
    
    def _set_connected(self, connected):
        """Actualiza estado de conexión."""
        if connected:
            self.status_lbl.configure(text="🟢 CONECTADO", text_color="#27ae60")
        else:
            self.token = None
            self.status_lbl.configure(text="⚫ DESCONECTADO", text_color="#e74c3c")
    
    def _on_mode_change(self, mode):
        """Cambia modo Manual/Excel."""
        if mode == "Manual":
            self.excel_btn.grid_forget()
            self.excel_lbl.grid_forget()
            self.docs_text.configure(state="normal")
        else:
            self.excel_btn.grid(row=0, column=2, padx=6, pady=4)
            self.excel_lbl.grid(row=0, column=3, padx=(0,6), pady=4, sticky="w")
            self.docs_text.configure(state="normal")
            self.docs_text.delete("1.0", "end")
            self.docs_text.configure(state="disabled")
        self._update_docs_count()
    
    def _update_docs_count(self, event=None):
        """Actualiza contador."""
        count = len(self._get_documents())
        self.docs_lbl.configure(text=f"📑 Documentos: {count}")
    
    def _select_excel(self):
        """Carga archivo Excel."""
        path = filedialog.askopenfilename(title="Seleccionar Excel", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        
        try:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            ws = wb.active
            
            existing = set(self.excel_docs)
            loaded, skipped = [], 0
            
            for row in ws.iter_rows(min_row=3, values_only=True):
                col_c = "" if len(row) < 3 or row[2] is None else str(row[2])
                col_d = "" if len(row) < 4 or row[3] is None else str(row[3])
                val = f"{col_c}{col_d}".strip()
                
                if val:
                    if val in existing:
                        skipped += 1
                    else:
                        loaded.append(val)
                        existing.add(val)
            
            wb.close()
            self.excel_docs.extend(loaded)
            self.excel_lbl.configure(text=f"📁 {os.path.basename(path)}", text_color="white")
            
            self.docs_text.configure(state="normal")
            self.docs_text.delete("1.0", "end")
            self.docs_text.insert("1.0", "\n".join(self.excel_docs[:500]))
            self.docs_text.configure(state="disabled")
            
            self._update_docs_count()
            messagebox.showinfo("Importado", f"✓ {len(loaded)} documentos.\n⚠️ {skipped} duplicados omitidos.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer:\n{e}")
    
    def _get_documents(self):
        """Obtiene lista de documentos."""
        if self.mode_selector.get() == "Manual":
            text = self.docs_text.get("1.0", "end").strip()
            return [l.strip() for l in text.splitlines() if l.strip()]
        return self.excel_docs.copy()
    
    def _do_clear(self):
        """Limpia todo."""
        self.docs_text.configure(state="normal")
        self.docs_text.delete("1.0", "end")
        if self.mode_selector.get() == "Excel":
            self.docs_text.configure(state="disabled")
        
        for item in self.table.get_children():
            self.table.delete(item)
        
        self.excel_docs.clear()
        self.excel_lbl.configure(text="", text_color="gray")
        self.results_count_lbl.configure(text="0 registros")
        self.progressbar.set(0)
        self.progress_lbl.configure(text="⏳ Esperando acción...", text_color="gray")
        self._update_docs_count()
    
    def _do_login(self):
        """Realiza login."""
        email = self.email_entry.get().strip()
        password = self.password_entry.get().strip()
        
        if not email or not password:
            messagebox.showwarning("Campos vacíos", "Ingresa correo y contraseña.")
            return
        
        self.login_btn.configure(state="disabled", text="🔄 Conectando...")
        self.progress_lbl.configure(text="🔄 Iniciando sesión...", text_color="orange")
        
        def task():
            try:
                res = requests.post(self.LOGIN_URL, json={"email": email, "password": password}, timeout=10)
                data = res.json()
                
                if "token" in data:
                    self.token = data["token"]
                    self.session.headers.update({"Authorization": f"Bearer {self.token}"})
                    self._save_token(self.token)
                    self.after(0, lambda: self._set_connected(True))
                    self.after(0, lambda: self.progress_lbl.configure(text="✅ Sesión iniciada.", text_color="#27ae60"))
                    self.after(0, lambda: messagebox.showinfo("Conectado", "✓ Sesión iniciada."))
                else:
                    self._delete_token()
                    self.after(0, lambda: self._set_connected(False))
                    self.after(0, lambda: self.progress_lbl.configure(text="❌ Credenciales inválidas.", text_color="#e74c3c"))
                    self.after(0, lambda: messagebox.showerror("Error", "Credenciales inválidas."))
            except Exception as e:
                self.after(0, lambda: self.progress_lbl.configure(text=f"❌ Error: {e}", text_color="#e74c3c"))
                self.after(0, lambda: messagebox.showerror("Error", f"Error de conexión:\n{e}"))
            finally:
                self.after(0, lambda: self.login_btn.configure(state="normal", text="🔌 Conectar"))
        
        threading.Thread(target=task, daemon=True).start()
    
    def _query_one(self, doc):
        """Consulta un afiliado."""
        if not self.token:
            return doc, "DESCONECTADO", "", ""
        
        doc_type = re.sub(r"\d", "", doc).upper()
        doc_num = re.sub(r"\D", "", doc)
        type_id = self.DOC_TYPES.get(doc_type)
        
        if not type_id or not doc_num:
            return doc, "FORMATO INVÁLIDO (Ej: CC1234567890)", "", ""
        
        try:
            res = self.session.get(f"{self.BASE_URL}/{doc_num}/{type_id}", timeout=10)
            
            if res.status_code in (401, 500):
                self.token = None
                self._delete_token()
                self.after(0, lambda: self._set_connected(False))
                return doc, "TOKEN EXPIRADO", "", ""
            
            if res.status_code == 200:
                data = res.json()
                
                if "error" in data:
                    return doc, data["error"], "", ""
                
                name = " ".join(filter(None, [
                    data.get("primer_nombre"), data.get("segundo_nombre"),
                    data.get("primer_apellido"), data.get("segundo_apellido")
                ])).strip() or "SIN NOMBRE"
                ips = data.get("ips", {}).get("nombre", "N/A")
                status = data.get("estado_afiliado", {}).get("nombre", "N/A")
                return doc, name, ips, status
            
            return doc, f"ERROR ({res.status_code}) - Usuario no encontrado", "", ""
        except Exception as e:
            return doc, f"ERROR: {e}", "", ""
    
    def _do_query(self):
        """Ejecuta consultas."""
        if not self.token:
            messagebox.showwarning("Sin conexión", "Inicia sesión primero.")
            return
        
        docs = self._get_documents()
        if not docs:
            messagebox.showwarning("Sin documentos", "Ingresa documentos.")
            return
        
        for item in self.table.get_children():
            self.table.delete(item)
        
        self._set_buttons_state("disabled")
        self.is_processing = True
        
        def task():
            total = len(docs)
            for idx, doc in enumerate(docs, 1):
                if not self.is_processing:
                    break
                
                result = self._query_one(doc)
                self.after(0, lambda r=result: self.table.insert("", "end", values=r))
                self.after(0, lambda i=idx, t=total: self._update_progress(i, t))
                time.sleep(random.uniform(1.5, 2.5))
            
            self.after(0, self._on_query_done)
        
        threading.Thread(target=task, daemon=True).start()
    
    def _update_progress(self, current, total):
        """Actualiza progreso."""
        pct = current / total
        self.progressbar.set(pct)
        self.progress_lbl.configure(text=f"⏳ Consultando... {current}/{total} ({int(pct*100)}%)", text_color="orange")
        self.results_count_lbl.configure(text=f"{current} registros")
    
    def _on_query_done(self):
        """Finaliza consultas."""
        self.is_processing = False
        self._set_buttons_state("normal")
        total = len(self.table.get_children())
        self.progressbar.set(1)
        self.progress_lbl.configure(text=f"✅ Completado. {total} registros.", text_color="#27ae60")
        messagebox.showinfo("Completado", f"✓ {total} registros consultados.")
    
    def _set_buttons_state(self, state):
        """Habilita/deshabilita botones."""
        self.query_btn.configure(state=state)
        self.export_btn.configure(state=state)
        self.clear_btn.configure(state=state)
    
    def _do_export(self):
        """Exporta a Excel."""
        if not self.table.get_children():
            messagebox.showwarning("Sin datos", "No hay datos.")
            return
        
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Afiliados"
            ws.append(["Tipo", "Número", "Nombre", "IPS", "Estado"])
            
            for item in self.table.get_children():
                doc, nombre, ips, estado = self.table.item(item, "values")
                tipo = re.sub(r"\d", "", doc).upper()
                numero = re.sub(r"\D", "", doc)
                ws.append([tipo, numero, nombre, ips, estado])
            
            wb.save(path)
            messagebox.showinfo("Exportado", f"✓ Guardado:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar:\n{e}")


if __name__ == "__main__":
    root = ctk.CTk()
    root.title("Horus Health - Consulta de Afiliados")
    root.geometry("1000x600")
    root.minsize(850, 500)
    app = HorusApp(root)
    app.pack(fill="both", expand=True)
    root.mainloop()