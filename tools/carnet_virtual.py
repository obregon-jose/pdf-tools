import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
from datetime import datetime
import re
import os
from threading import Thread
from tools.vaccine_catalog import VACCINE_CATALOG

# Nombres de meses en español
MESES_ESP = {
    1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DIC"
}

# Columnas que deben ser numéricas en el Excel de salida
NUMERIC_COLUMNS = [
    'documento',
    'telefono',
    'c_desplazamiento',
    'c_discapacidad',
    'c_usuaria',
    'dosis_aplicada',
    'lote_jeringa',
    'evento_postvacunal'
]

# Diccionario para normalización del nombre del brazo
ARM_DICTIONARY = {
    # Left arm variations / Variaciones brazo izquierdo
    'IZQUIERDO': 'IZQUIERDO',
    'IZQUIERDA': 'IZQUIERDO',
    'IZQ': 'IZQUIERDO',
    'IZ': 'IZQUIERDO',
    'I': 'IZQUIERDO',
    'IQ': 'IZQUIERDO',
    
    # Right arm variations / Variaciones brazo derecho
    'DERECHO': 'DERECHO',
    'DERECHA': 'DERECHO',
    'DER': 'DERECHO',
    'DE': 'DERECHO',
    'D': 'DERECHO',
    'DR': 'DERECHO',
}


class CarnetVirtualApp(ctk.CTkFrame):
    
    def __init__(self, master, go_home=None):
        super().__init__(master)
        
        # ===== ESTADO DE LA APLICACIÓN =====
        self.input_file_path = ""
        self.output_folder_path = ""
        
        # Diccionarios dinámicos para lotes y vencimientos
        self.lotes = {}
        self.vencimientos = {}
        
        # Contadores dinámicos para estadísticas
        self.vaccine_counts = {}
        
        # Vacunas detectadas en el archivo
        self.vaccines_detected = []
        
        # Historial para deshacer
        self.last_operation = None
        
        # Widgets dinámicos por vacuna
        self.vaccine_widgets = {}
        self.vaccine_entries = {}
        
        # ===== CONSTRUIR UI =====
        self._create_widgets()
    
    # ==================== CONSTRUCCIÓN DE UI ====================
    
    def _create_widgets(self):
        """Crea todos los widgets de la interfaz siguiendo el diseño del modelo."""
        
        # ===== FRAME SUPERIOR (compacto) =====
        self.top_frame = ctk.CTkFrame(self)
        self.top_frame.pack(fill="x", padx=6, pady=6)
        
        # Fila 0: Archivo de entrada
        lbl_input = ctk.CTkLabel(self.top_frame, text="Entrada:", width=60, anchor="w")
        lbl_input.grid(row=0, column=0, padx=(6, 4), pady=4, sticky="w")
        
        self.input_entry = ctk.CTkEntry(
            self.top_frame,
            placeholder_text="Selecciona el archivo Excel de origen",
            width=500
        )
        self.input_entry.grid(row=0, column=1, padx=(0, 6), pady=4, sticky="we")
        self.input_entry.bind("<Return>", lambda e: self._load_from_entry())
        
        btn_select_input = ctk.CTkButton(
            self.top_frame,
            text="Seleccionar Excel",
            width=140,
            command=self._on_select_input
        )
        btn_select_input.grid(row=0, column=2, padx=(0, 6), pady=4)
        
        # Fila 1: Carpeta de salida
        lbl_output = ctk.CTkLabel(self.top_frame, text="Salida:", width=60, anchor="w")
        lbl_output.grid(row=1, column=0, padx=(6, 4), pady=4, sticky="w")
        
        self.output_entry = ctk.CTkEntry(
            self.top_frame,
            placeholder_text="Carpeta de salida (opcional)",
            width=500
        )
        self.output_entry.grid(row=1, column=1, padx=(0, 6), pady=4, sticky="we")
        
        btn_select_output = ctk.CTkButton(
            self.top_frame,
            text="Seleccionar Carpeta",
            width=140,
            command=self._on_select_output
        )
        btn_select_output.grid(row=1, column=2, padx=(0, 6), pady=4)
        
        # El resto de la UI se creará dinámicamente después de cargar el archivo
        
        self.top_frame.grid_columnconfigure(1, weight=1)
        
        # ===== PANEL DE LOG =====
        log_frame = ctk.CTkFrame(self)
        log_frame.pack(fill="both", expand=True, padx=6, pady=6)
        
        # Nota informativa
        # self.info_note = ctk.CTkLabel(
        #     log_frame,
        #     text="Nota: Selecciona el archivo Excel para detectar automáticamente las vacunas a procesar",
        #     text_color="gray",
        #     font=("Arial", 12, "bold")
        # )
        # self.info_note.pack(anchor="w", padx=4, pady=(2, 4))
        
        # Textbox para el log
        self.log_text = ctk.CTkTextbox(
            log_frame,
            height=200,
            font=ctk.CTkFont(family="Consolas", size=11)
        )
        self.log_text.pack(fill="both", expand=True, padx=4, pady=2)
        
        # ===== PANEL DE PROGRESO (DEBAJO DEL LOG) =====
        progress_frame = ctk.CTkFrame(self)
        progress_frame.pack(fill="x", padx=6, pady=6)
        
        self.progress_bar = ctk.CTkProgressBar(progress_frame, height=15)
        self.progress_bar.pack(fill="x", padx=15, pady=(10, 5))
        self.progress_bar.set(0)
        
        # self.status_label = ctk.CTkLabel(
        #     progress_frame,
        #     text="⏳ Esperando datos para procesar...",
        #     text_color="gray",
        #     anchor="w"
        # )
        # self.status_label.pack(anchor="w", padx=15, pady=(0, 10))
        
        self._log_message("✅ Sistema iniciado. Listo para procesar datos de vacunación. ")
        self._log_message("💉🏥 Seleccione una jornada para Generar su base de CARNET VIRTUAL.")
    
    def _create_vaccine_fields(self):
        """Crea dinámicamente los campos de vacunas según lo detectado en el archivo."""
        # Limpiar widgets previos
        for vaccine_id in self.vaccine_widgets:
            for widget in self.vaccine_widgets[vaccine_id]:
                widget.destroy()
        
        self.vaccine_widgets.clear()
        self.vaccine_entries.clear()
        
        current_row = 2
        
        # Espaciador y título de biológicos
        if self.vaccines_detected:
            spacer = ctk.CTkFrame(self.top_frame, height=2)
            spacer.grid(row=current_row, column=0, columnspan=3)
            
            if self.vaccines_detected[0] not in self.vaccine_widgets:
                self.vaccine_widgets[self.vaccines_detected[0]] = []
            self.vaccine_widgets[self.vaccines_detected[0]].append(spacer)
            
            current_row += 1
        
        # Crear campos para cada vacuna detectada
        for vaccine_id in self.vaccines_detected:
            vaccine_info = VACCINE_CATALOG.get(vaccine_id)
            
            if not vaccine_info:
                continue
            
            # Inicializar lista de widgets para esta vacuna
            self.vaccine_widgets[vaccine_id] = []
            self.vaccine_entries[vaccine_id] = {}
            
            # Label de la vacuna
            lbl_vaccine = ctk.CTkLabel(
                self.top_frame,
                text=f"{vaccine_info['keywords'][0]}:",
                width=100,
                anchor="w",
                font=ctk.CTkFont(weight="bold")
            )
            lbl_vaccine.grid(row=current_row, column=0, padx=(6, 4), pady=4, sticky="w")
            self.vaccine_widgets[vaccine_id].append(lbl_vaccine)
            
            # Container para lote y vencimiento HORIZONTALMENTE
            vaccine_container = ctk.CTkFrame(self.top_frame, fg_color="transparent")
            vaccine_container.grid(row=current_row, column=1, columnspan=2, padx=(0, 6), pady=4, sticky="we")
            self.vaccine_widgets[vaccine_id].append(vaccine_container)
            
            # Lote (izquierda)
            lbl_lote = ctk.CTkLabel(vaccine_container, text="Lote:", anchor="w", width=50)
            lbl_lote.pack(side="left", padx=(0, 4))
            
            entry_lote = ctk.CTkEntry(
                vaccine_container,
                placeholder_text=f"Ej: {vaccine_info.get('lote_ejemplo', 'ABC123')}",
                width=150
            )
            entry_lote.pack(side="left", padx=(0, 20))
            self.vaccine_entries[vaccine_id]['lote'] = entry_lote
            
            # Vencimiento (derecha)
            lbl_venc = ctk.CTkLabel(vaccine_container, text="Vencimiento:", anchor="w", width=90)
            lbl_venc.pack(side="left", padx=(0, 4))
            
            entry_venc = ctk.CTkEntry(
                vaccine_container,
                placeholder_text="DD/MM/AAAA o DD MM AAAA",
                width=200
            )
            entry_venc.pack(side="left")
            entry_venc.bind("<FocusOut>", lambda e, ent=entry_venc: self._validate_date_field(ent))
            self.vaccine_entries[vaccine_id]['vencimiento'] = entry_venc
            
            current_row += 1
        
        # Espaciador
        spacer = ctk.CTkFrame(self.top_frame, height=2)
        spacer.grid(row=current_row, column=0, columnspan=3)
        if self.vaccines_detected:
            self.vaccine_widgets[self.vaccines_detected[0]].append(spacer)
        current_row += 1
        
        # Botones de acción
        btn_frame = ctk.CTkFrame(self.top_frame)
        btn_frame.grid(row=current_row, column=0, columnspan=3, sticky="we", padx=6, pady=(0, 4))
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        if self.vaccines_detected:
            self.vaccine_widgets[self.vaccines_detected[0]].append(btn_frame)
        
        self.process_button = ctk.CTkButton(
            btn_frame,
            text="🚀 Procesar Datos",
            command=self._on_process,
            fg_color="#1f6feb"
        )
        self.process_button.grid(row=0, column=0, padx=6, pady=4, sticky="we")
        
        self.undo_button = ctk.CTkButton(
            btn_frame,
            text="↶ Deshacer",
            command=self._on_undo,
            fg_color="#f0ad4e",
            text_color="black"
        )
        self.undo_button.grid(row=0, column=1, padx=6, pady=4, sticky="we")
        self.undo_button.configure(state="disabled")
    
    # ==================== DETECCIÓN DE VACUNAS ====================
    
    def _identify_vaccine(self, vaccine_text):
        """Identifica la vacuna a partir del texto usando el catálogo."""
        vaccine_text_upper = self._clean_text(vaccine_text)
        
        for vaccine_id, vaccine_info in VACCINE_CATALOG.items():
            for keyword in vaccine_info['keywords']:
                if keyword in vaccine_text_upper:
                    return vaccine_id
        
        return 'UNKNOWN'
    
    def _detect_vaccines_in_file(self, file_path):
        """Detecta qué vacunas están presentes en el archivo Excel."""
        try:
            df = pd.read_excel(file_path)
            
            # Log de columnas encontradas para debugging
            # self._log_message(f"📋 Columnas del archivo: {len(df.columns)}")
            
            # Buscar columna de vacunas
            column_mapping = self._map_columns(df.columns.tolist())
            
            if 'vacuna' not in column_mapping:
                self._log_message("⚠️ No se encontró columna de vacunas")
                # Mostrar columnas disponibles para ayudar al usuario
                self._log_message(f"   Columnas disponibles: {', '.join(df.columns.tolist()[:10])}...")
                return
            
            vacuna_col = column_mapping['vacuna']
            
            # Revisar todas las vacunas en el archivo
            vaccines_found = set()
            
            for value in df[vacuna_col].dropna():
                # Parsear el texto para obtener TODAS las vacunas de la celda
                biologicos = self._parse_biologicos(str(value))
                
                for biologico in biologicos:
                    vaccine_id = self._identify_vaccine(biologico)
                    if vaccine_id != 'UNKNOWN':
                        vaccines_found.add(vaccine_id)
            
            # Ordenar por prioridad
            self.vaccines_detected = sorted(list(vaccines_found))
            
            # Inicializar contadores
            self.vaccine_counts = {v_id: 0 for v_id in self.vaccines_detected}
            
            # Log de detección
            if self.vaccines_detected:
                vaccine_names = [
                    VACCINE_CATALOG[v_id]['display_name'] 
                    for v_id in self.vaccines_detected
                ]
                self._log_message(f"🔍 Vacunas detectadas: {', '.join(vaccine_names)}")
                
                # Verificar columnas de brazo detectadas
                for v_id in self.vaccines_detected:
                    brazo_key = f'brazo_{v_id.lower()}'
                    if brazo_key in column_mapping:
                        pass
                        # self._log_message(f"   ✅ Columna brazo {VACCINE_CATALOG[v_id]['display_name']}: '{column_mapping[brazo_key]}'")
                    else:
                        self._log_message(f"   ⚠️ Sin columna brazo para {VACCINE_CATALOG[v_id]['display_name']} (usará: {VACCINE_CATALOG[v_id]['default_arm']})")
            else:
                self._log_message("⚠️ No se detectaron vacunas conocidas en el archivo")
            
        except Exception as e:
            self._log_message(f"⚠️ Error al detectar vacunas: {str(e)}")
            self.vaccines_detected = []
    # ==================== MANEJADORES DE EVENTOS ====================
    
    def _on_select_input(self):
        """Abre diálogo para seleccionar archivo Excel de entrada."""
        file_path = filedialog.askopenfilename(
            title="Selecciona el archivo Excel de entrada",
            filetypes=[
                ("Archivos Excel", "*.xlsx *.xls"),
                ("Todos los archivos", "*.*")
            ]
        )
        
        if file_path:
            # Limpiar log anterior
            self._clear_log()
            
            self.input_entry.delete(0, "end")
            self.input_entry.insert(0, file_path)
            self.input_file_path = file_path
            self._log_message(f"📂 Archivo de entrada: {os.path.basename(file_path)}")
            
            # Detectar vacunas en el archivo
            self._detect_vaccines_in_file(file_path)
            
            # Crear campos dinámicos
            self._create_vaccine_fields()
            
            # Auto-llenar carpeta de salida si está vacía
            if not self.output_entry.get():
                output_dir = os.path.dirname(file_path)
                self.output_entry.delete(0, "end")
                self.output_entry.insert(0, output_dir)
                self.output_folder_path = output_dir
    
    def _load_from_entry(self):
        """Carga archivo desde la ruta escrita."""
        typed_path = self.input_entry.get().strip()
        if typed_path:
            if os.path.exists(typed_path) and os.path.isfile(typed_path):
                # Limpiar log anterior
                self._clear_log()
                
                self.input_file_path = typed_path
                self._log_message(f"📂 Archivo cargado: {os.path.basename(typed_path)}")
                
                # Detectar vacunas
                self._detect_vaccines_in_file(typed_path)
                
                # Crear campos dinámicos
                self._create_vaccine_fields()
            else:
                messagebox.showerror("Error", f"El archivo no existe:\n{typed_path}")
       
    def _on_select_output(self):
        """Abre diálogo para seleccionar carpeta de salida."""
        folder = filedialog.askdirectory(title="Selecciona carpeta de salida")
        
        if folder:
            self.output_entry.delete(0, "end")
            self.output_entry.insert(0, folder)
            self.output_folder_path = folder
            self._log_message(f"💾 Carpeta de salida: {folder}")
    
    def _on_process(self):
        """Inicia el procesamiento de datos."""
        # Validar entradas
        errors = self._validate_inputs()
        
        if errors:
            error_message = "Por favor corrija los siguientes errores:\n\n" + "\n".join(errors)
            messagebox.showwarning("Advertencia", error_message)
            self._log_message("❌ Validación fallida: Faltan campos requeridos")
            return
        
        # Capturar valores de lotes y vencimientos
        for vaccine_id in self.vaccines_detected:
            lote_text = self.vaccine_entries[vaccine_id]['lote'].get().strip().upper()
            
            # Eliminar todos los espacios del lote
            lote_text = lote_text.replace(" ", "")
            
            venc_text = self.vaccine_entries[vaccine_id]['vencimiento'].get().strip()
            
            venc_parsed = self._parse_and_format_date(venc_text)
            
            if not venc_parsed:
                vaccine_name = VACCINE_CATALOG[vaccine_id]['display_name']
                messagebox.showerror("Error de Fecha", f"La fecha de vencimiento de {vaccine_name} no es válida.")
                self._log_message(f"❌ Error: Fecha de vencimiento {vaccine_name} inválida")
                return
            
            self.lotes[vaccine_id] = lote_text
            self.vencimientos[vaccine_id] = venc_parsed
        
        # Reiniciar contadores
        for vaccine_id in self.vaccines_detected:
            self.vaccine_counts[vaccine_id] = 0
        
        # Deshabilitar botón
        self.process_button.configure(state="disabled", text="⏳ Procesando...")
        
        # Ejecutar en hilo separado
        thread = Thread(target=self._process_data)
        thread.start()
    
    def _on_undo(self):
        """Deshace la última operación eliminando el archivo creado."""
        if not self.last_operation:
            messagebox.showinfo("Info", "No hay operación para deshacer.")
            return
        
        created_file = self.last_operation.get('created')
        
        if not created_file or not os.path.exists(created_file):
            messagebox.showwarning("Advertencia", "El archivo creado ya no existe.")
            self.last_operation = None
            self.undo_button.configure(state="disabled")
            return
        
        response = messagebox.askyesno(
            "Confirmar Deshacer",
            f"¿Eliminar el archivo creado?\n\n{created_file}"
        )
        
        if not response:
            return
        
        try:
            os.remove(created_file)
            messagebox.showinfo("Deshacer completado", f"Archivo eliminado:\n{created_file}")
            self._log_message(f"🗑️ Archivo eliminado: {os.path.basename(created_file)}")
            self.last_operation = None
            self.undo_button.configure(state="disabled")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar:\n{str(e)}")
    
    # ==================== VALIDACIONES ====================
    
    def _validate_inputs(self):
        """Valida todas las entradas requeridas."""
        errors = []
        
        if not self.input_entry.get().strip():
            errors.append("• Debe seleccionar un archivo de entrada")
        
        if not self.vaccines_detected:
            errors.append("• No se detectaron vacunas en el archivo")
        
        for vaccine_id in self.vaccines_detected:
            vaccine_name = VACCINE_CATALOG[vaccine_id]['display_name']
            
            if not self.vaccine_entries[vaccine_id]['lote'].get().strip():
                errors.append(f"• Debe ingresar el lote de la {vaccine_name}")
            
            if not self.vaccine_entries[vaccine_id]['vencimiento'].get().strip():
                errors.append(f"• Debe ingresar el vencimiento de la {vaccine_name}")
        
        return errors
    
    def _validate_date_field(self, entry_widget):
        """Valida y formatea automáticamente el campo de fecha."""
        date_text = entry_widget.get().strip()
        
        if not date_text:
            return
        
        formatted_date = self._parse_and_format_date(date_text)
        
        if formatted_date:
            entry_widget.delete(0, "end")
            entry_widget.insert(0, formatted_date)
            entry_widget.configure(border_color=("gray50", "gray75"))
        else:
            entry_widget.configure(border_color="red")
            self._log_message(f"⚠️ Fecha inválida detectada: {date_text}")
    
    def _parse_and_format_date(self, date_text):
        """Parsea y formatea una fecha en múltiples formatos aceptados."""
        if not date_text:
            return None
        
        date_text = date_text.strip()
        normalized = re.sub(r'[/\-\.]', ' ', date_text)
        
        formats_to_try = [
            '%d %m %Y',
            '%d/%m/%Y',
            '%d-%m-%Y',
            '%d %m %y',
            '%d/%m/%y',
            '%d-%m-%y',
        ]
        
        try:
            parsed_date = datetime.strptime(normalized, '%d %m %Y')
            return parsed_date.strftime('%d/%m/%Y')
        except ValueError:
            pass
        
        try:
            parsed_date = datetime.strptime(normalized, '%d %m %y')
            return parsed_date.strftime('%d/%m/%Y')
        except ValueError:
            pass
        
        for fmt in formats_to_try:
            try:
                parsed_date = datetime.strptime(date_text, fmt)
                return parsed_date.strftime('%d/%m/%Y')
            except ValueError:
                continue
        
        return None
    
    # ==================== UTILIDADES DE UI ====================
    
    def _log_message(self, message):
        """Agrega un mensaje al registro."""
        self.log_text.insert("end", f"{message}\n")
        self.log_text.see("end")
    
    def _update_progress(self, value, status_text):
        """Actualiza la barra de progreso y etiqueta de estado."""
        self.progress_bar.set(value)
        # self.status_label.configure(text=status_text)

    def _clear_log(self):
        """
        Limpia el contenido del log.
        Clears the log content.
        """
        self.log_text.delete("1.0", "end")

    # ==================== PROCESAMIENTO DE DATOS ====================
    
    def _process_data(self):
        """Procesa los datos de vacunación."""
        try:
            self._log_message("🚀 Iniciando procesamiento de datos...")
            self._update_progress(0.1, "📂 Leyendo archivo de entrada...")
            
            df_input = pd.read_excel(self.input_file_path)
            total_rows = len(df_input)
            self._log_message(f"✅ Archivo leído correctamente. Total de registros: {total_rows}")
            
            self._update_progress(0.2, "🔍 Mapeando columnas...")
            
            column_mapping = self._map_columns(df_input.columns.tolist())
            self._log_message(f"📋 Columnas mapeadas: {len(column_mapping)}")
            
            first_date = None
            if 'fecha_atencion' in column_mapping and len(df_input) > 0:
                first_date = df_input.iloc[0].get(column_mapping['fecha_atencion'], None)
                self._log_message(f"📅 Fecha identificada: {first_date}")
            
            self._update_progress(0.3, "🔄 Procesando registros...")
            
            processed_rows = []
            
            for index, row in df_input.iterrows():
                progress = 0.3 + (0.5 * (index + 1) / total_rows)
                self._update_progress(progress, f"🔄 Procesando registro {index + 1} de {total_rows}...")
                
                rows = self._process_row(row, column_mapping)
                processed_rows.extend(rows)
            
            self._log_message(f"✅ Registros procesados: {len(processed_rows)}")
            
            self._update_progress(0.85, "💾 Generando archivo de salida...")
            
            output_path = self._generate_output_filename(
                self.input_file_path,
                first_date
            )
            self._log_message(f"📁 Generando: {os.path.basename(output_path)}")
            
            output_columns = [
                'fecha_vacunacion', 'tipo_documento', 'documento', 'fecha_nacimiento',
                'sexo', 'primer_apellido', 'segundo_apellido', 'nombres', 'regimen',
                'aseguradora', 'municipio', 'area_residencia', 'barrio', 'direccion',
                'eps', 'telefono', 'grupo_etnico', 'c_desplazamiento', 'c_discapacidad',
                'correo', 'c_usuaria', 'fecha_parto', 'tipo_poblacion', 'dosis_aplicada',
                'biologico', 'lote_biologico', 'jeringa', 'lote_jeringa',
                'evento_postvacunal', 'vacunador', 'municipio_reporta', 'novedad',
                'desc_novedad', 'modalidad_vacunacion', 'nota_enfermeria', 'jornada'
            ]
            
            df_output = pd.DataFrame(processed_rows, columns=output_columns)
            
            self._create_excel_file(df_output, output_path)
            
            self._update_progress(1.0, "✅ ¡Proceso completado exitosamente!")
            
            # Guardar para deshacer
            self.last_operation = {
                'created': output_path,
                'total_rows': total_rows,
                'total_applications': len(df_output)
            }
            self.undo_button.configure(state="normal")
            
            # Mostrar resumen
            self._log_message("=" * 60)
            self._log_message("📊 RESUMEN DEL PROCESO:")
            self._log_message(f"   • Pacientes procesados: {total_rows}")
            self._log_message(f"   • Total aplicaciones: {len(df_output)}")
            
            summary_parts = []
            for vaccine_id, count in self.vaccine_counts.items():
                if count > 0:
                    vaccine_name = VACCINE_CATALOG[vaccine_id]['keywords'][0]
                    self._log_message(f"   • {vaccine_name}: {count}")
                    summary_parts.append(f"{vaccine_name}: {count}")
            
            self._log_message(f"   • Archivo generado: {os.path.basename(output_path)}")
            self._log_message("=" * 60)
            self._log_message("✅ ¡PROCESO COMPLETADO EXITOSAMENTE!")
            
            # Mensaje de resumen
            summary_text = "\n   • ".join(summary_parts) if summary_parts else "Sin vacunas procesadas"
            
            self.after(0, lambda: messagebox.showinfo(
                "Proceso Completado",
                f"¡El archivo se ha generado correctamente!\n\n"
                f"👥 Pacientes procesados: {total_rows}\n"
                f"💉 Total aplicaciones: {len(df_output)}\n\n"
                f"   • {summary_text}\n\n"
                f"📁 Archivo:\n{os.path.basename(output_path)}\n\n"
                f"📍 Ubicación:\n{os.path.dirname(output_path)}"
            ))
            
        except FileNotFoundError:
            self._log_message(f"❌ Error: No se encontró el archivo de entrada")
            self.after(0, lambda: messagebox.showerror("Error", "No se encontró el archivo de entrada"))
        except Exception as e:
            self._log_message(f"❌ Error durante el procesamiento: {str(e)}")
            self.after(0, lambda: messagebox.showerror("Error", f"Error durante el procesamiento:\n{str(e)}"))
        finally:
            self.after(0, lambda: self.process_button.configure(
                state="normal",
                text="🚀 Procesar Datos"
            ))
    
    def _generate_output_filename(self, input_path, date_from_data):
        """Genera el nombre del archivo de salida basado en la fecha y vacunas detectadas."""
        output_dir = self.output_entry.get().strip()
        if not output_dir:
            output_dir = os.path.dirname(input_path)
        
        try:
            if isinstance(date_from_data, datetime):
                day = date_from_data.day
                month = date_from_data.month
            elif isinstance(date_from_data, str):
                date_from_data = str(date_from_data).strip()
                
                for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d %m %Y']:
                    try:
                        parsed_date = datetime.strptime(date_from_data, fmt)
                        day = parsed_date.day
                        month = parsed_date.month
                        break
                    except ValueError:
                        continue
                else:
                    day = datetime.now().day
                    month = datetime.now().month
            else:
                day = datetime.now().day
                month = datetime.now().month
        except Exception:
            day = datetime.now().day
            month = datetime.now().month
        
        month_name = MESES_ESP.get(month, "MES")
        
        # Generar nombre según vacunas detectadas
        vaccine_names = [
            VACCINE_CATALOG[v_id]['keywords'][0] 
            for v_id in self.vaccines_detected
        ]
        
        vaccine_str = " - ".join(vaccine_names) if vaccine_names else "VACUNAS"
        
        filename = f"BASE CARNET VIRTUAL [{vaccine_str}] DIA {day} {month_name}.xlsx"
        full_path = os.path.join(output_dir, filename)
        
        return full_path
    
    def _create_excel_file(self, df, output_path):
        """Crea archivo Excel con formato correcto."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        
        self._log_message("📝 Creando archivo Excel...")
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Datos Vacunación"
        
        header_font = Font(name='Calibri', size=11, bold=True)
        header_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        
        cell_font = Font(name='Calibri', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        numeric_col_indices = set()
        for idx, header in enumerate(headers, 1):
            if header in NUMERIC_COLUMNS:
                numeric_col_indices.add(idx)
        
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                if col_idx in numeric_col_indices and value:
                    try:
                        numeric_value = int(re.sub(r'[^\d]', '', str(value)))
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=numeric_value)
                    except (ValueError, TypeError):
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                else:
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        column_widths = {
            'fecha_vacunacion': 15, 'tipo_documento': 12, 'documento': 15,
            'fecha_nacimiento': 15, 'sexo': 12, 'primer_apellido': 18,
            'segundo_apellido': 18, 'nombres': 25, 'regimen': 15,
            'aseguradora': 20, 'municipio': 12, 'area_residencia': 12,
            'barrio': 20, 'direccion': 30, 'eps': 20, 'telefono': 15,
            'grupo_etnico': 12, 'c_desplazamiento': 14, 'c_discapacidad': 14,
            'correo': 35, 'c_usuaria': 10, 'fecha_parto': 12,
            'tipo_poblacion': 12, 'dosis_aplicada': 12, 'biologico': 20,
            'lote_biologico': 15, 'jeringa': 20, 'lote_jeringa': 12,
            'evento_postvacunal': 16, 'vacunador': 15, 'municipio_reporta': 15,
            'novedad': 12, 'desc_novedad': 15, 'modalidad_vacunacion': 18,
            'nota_enfermeria': 80, 'jornada': 25
        }
        
        for col_idx, header in enumerate(headers, 1):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            width = column_widths.get(header, 15)
            worksheet.column_dimensions[col_letter].width = width
        
        worksheet.freeze_panes = 'A2'
        
        workbook.save(output_path)
        
        self._log_message(f"✅ Archivo Excel creado: {os.path.basename(output_path)}")
    
    # ==================== PROCESAMIENTO DE FILAS ====================
    
    def _map_columns(self, input_columns):
        """Mapea nombres de columnas de entrada a nombres esperados."""
        column_mapping = {}
        
        for col in input_columns:
            col_upper = str(col).upper().strip()
            
            if 'FECHA' in col_upper and 'ATENCION' in col_upper:
                column_mapping['fecha_atencion'] = col
            elif 'TIPO' in col_upper and 'DOCUMENTO' in col_upper:
                column_mapping['tipo_documento'] = col
            elif 'NUMERO' in col_upper and 'DOCUMENTO' in col_upper:
                column_mapping['numero_documento'] = col
            elif 'PRIMER NOMBRE' in col_upper or col_upper == 'PRIMER NOMBRE':
                column_mapping['primer_nombre'] = col
            elif 'SEGUNDO NOMBRE' in col_upper or col_upper == 'SEGUNDO NOMBRE':
                column_mapping['segundo_nombre'] = col
            elif 'PRIMER APELLIDO' in col_upper or col_upper == 'PRIMER APELLIDO':
                column_mapping['primer_apellido'] = col
            elif 'SEGUNDO APELLIDO' in col_upper or col_upper == 'SEGUNDO APELLIDO':
                column_mapping['segundo_apellido'] = col
            elif 'FECHA' in col_upper and 'NACIMIENTO' in col_upper:
                column_mapping['fecha_nacimiento'] = col
            elif 'TELEFON' in col_upper or 'NUMERO TELEFONICO' in col_upper:
                column_mapping['telefono'] = col
            elif 'CORREO' in col_upper:
                column_mapping['correo'] = col
            elif 'DIRECC' in col_upper:
                column_mapping['direccion'] = col
            elif 'BARRIO' in col_upper:
                column_mapping['barrio'] = col
            elif col_upper == 'EPS':
                column_mapping['eps'] = col
            elif 'SEXO' in col_upper:
                column_mapping['sexo'] = col
            elif 'VACUNA' in col_upper and 'NOMBRE' in col_upper:
                column_mapping['vacuna'] = col
            elif 'JORNADA' in col_upper or 'LUGAR' in col_upper:
                column_mapping['jornada'] = col
        
        # Mapear columnas de brazo dinámicamente buscando coincidencias con keywords del catálogo
        for col in input_columns:
            col_upper = str(col).upper().strip()
            
            # Buscar en cada vacuna del catálogo
            for vaccine_id, vaccine_info in VACCINE_CATALOG.items():
                # Ya encontramos esta columna de brazo, saltar
                if f'brazo_{vaccine_id.lower()}' in column_mapping:
                    continue
                
                # Buscar coincidencia con cualquier keyword de la vacuna
                for keyword in vaccine_info['keywords']:
                    # La columna contiene el keyword de la vacuna
                    if keyword in col_upper:
                        # Verificar que NO sea la columna principal de vacunas
                        if 'NOMBRE' not in col_upper and 'VACUNA' not in col_upper:
                            column_mapping[f'brazo_{vaccine_id.lower()}'] = col
                            # self._log_message(f"   📍 Columna brazo detectada: '{col}' → {vaccine_info['display_name']}")
                            break
                
                # Si ya encontramos, salir del loop de keywords
                if f'brazo_{vaccine_id.lower()}' in column_mapping:
                    break
        
        return column_mapping
    
    def _process_row(self, row, column_mapping):
        """Procesa una fila y retorna una o más filas de salida."""
        vacuna_texto = row.get(column_mapping.get('vacuna', ''), '')
        biologicos = self._parse_biologicos(vacuna_texto)
        
        if not biologicos:
            biologicos = ['SIN ESPECIFICAR']
        
        # Datos comunes del paciente
        fecha_vac = self._format_date(row.get(column_mapping.get('fecha_atencion', ''), ''))
        tipo_doc = self._clean_text(row.get(column_mapping.get('tipo_documento', ''), ''))
        documento = self._clean_document(row.get(column_mapping.get('numero_documento', ''), ''))
        fecha_nac = self._format_date(row.get(column_mapping.get('fecha_nacimiento', ''), ''))
        
        sexo_raw = row.get(column_mapping.get('sexo', ''), '')
        sexo = self._translate_sex(sexo_raw)
        
        primer_apellido = self._clean_text(row.get(column_mapping.get('primer_apellido', ''), ''))
        segundo_apellido = self._clean_text(row.get(column_mapping.get('segundo_apellido', ''), ''))
        nombres = self._get_full_name(
            row.get(column_mapping.get('primer_nombre', ''), ''),
            row.get(column_mapping.get('segundo_nombre', ''), '')
        )
        barrio = self._clean_text(row.get(column_mapping.get('barrio', ''), ''))
        direccion = self._clean_text(row.get(column_mapping.get('direccion', ''), ''))
        eps = self._clean_text(row.get(column_mapping.get('eps', ''), ''))
        telefono = self._clean_document(row.get(column_mapping.get('telefono', ''), ''))
        correo = self._clean_text(row.get(column_mapping.get('correo', ''), ''))
        jornada = self._clean_text(row.get(column_mapping.get('jornada', ''), ''))
        
        output_rows = []
        
        for biologico_raw in biologicos:
            biologico_raw = self._clean_text(biologico_raw)
            vaccine_id = self._identify_vaccine(biologico_raw)
            
            if vaccine_id == 'UNKNOWN':
                continue
            
            vaccine_info = VACCINE_CATALOG.get(vaccine_id)
            
            if not vaccine_info:
                continue
            
            # Obtener lote y vencimiento
            lote = self.lotes.get(vaccine_id, '')
            vencimiento = self.vencimientos.get(vaccine_id, '')
            
            # Obtener brazo
            brazo_col = column_mapping.get(f'brazo_{vaccine_id.lower()}')
            if brazo_col:
                brazo_raw = row.get(brazo_col, vaccine_info['default_arm'])
                brazo = self._normalize_arm(brazo_raw)
            else:
                brazo = vaccine_info['default_arm']
            
            # Nombre del biológico
            biologico_display = vaccine_info['display_name']
            
            # Jeringa
            jeringa = vaccine_info.get('jeringa', 'JERINGA PRELLENADA')
            
            # Incrementar contador
            self.vaccine_counts[vaccine_id] += 1
            
            # Generar nota de enfermería
            nota = self._generate_nursing_note(vaccine_id, lote, vencimiento, brazo)
            
            output_row = {
                'fecha_vacunacion': fecha_vac,
                'tipo_documento': tipo_doc,
                'documento': documento,
                'fecha_nacimiento': fecha_nac,
                'sexo': sexo,
                'primer_apellido': primer_apellido,
                'segundo_apellido': segundo_apellido,
                'nombres': nombres,
                'regimen': '',
                'aseguradora': '',
                'municipio': 'CALI',
                'area_residencia': 'URBANA',
                'barrio': barrio,
                'direccion': direccion,
                'eps': eps,
                'telefono': telefono,
                'grupo_etnico': 'NINGUNO',
                'c_desplazamiento': '0',
                'c_discapacidad': '0',
                'correo': correo,
                'c_usuaria': '0',
                'fecha_parto': 'NINGUNO',
                'tipo_poblacion': 'ADULTO',
                'dosis_aplicada': '1',
                'biologico': biologico_display,
                'lote_biologico': lote,
                'jeringa': jeringa,
                'lote_jeringa': '0',
                'evento_postvacunal': '0',
                'vacunador': '* AUXILIAR',
                'municipio_reporta': 'CALI',
                'novedad': 'NINGUNO',
                'desc_novedad': 'SIN NOVEDAD',
                'modalidad_vacunacion': 'EXTRAMURAL',
                'nota_enfermeria': nota,
                'jornada': jornada
            }
            
            output_rows.append(output_row)
        
        return output_rows
    
    def _generate_nursing_note(self, vaccine_id, lote, vencimiento, brazo):
        """Genera la nota de enfermería usando información del catálogo."""
        vaccine_info = VACCINE_CATALOG.get(vaccine_id, {})
        vaccine_desc = vaccine_info.get('description', 'vacuna no especificada')
        brazo_text = brazo.upper() if brazo else "IZQUIERDO"
        
        note = (
            f"Asiste usuario para aplicación de vacuna contra {vaccine_desc}, "
            f"jeringa prellenada, lote {lote}, con fecha de vencimiento {vencimiento}. "
            f"Se brinda explicación del procedimiento y se informa sobre posibles efectos posvacunales. "
            f"Se realiza administración del biológico en región deltoides de brazo {brazo_text.lower()} sin incidencias. "
            f"Se entrega carnet de vacunas debidamente diligenciado. "
            f"Paciente refiere haber comprendido la información proporcionada."
        )
        
        return note
    
    # ==================== UTILIDADES DE PROCESAMIENTO ====================
    
    def _clean_text(self, text):
        """Elimina espacios extra y convierte a mayúsculas."""
        if pd.isna(text) or text is None:
            return ""
        text = str(text).strip()
        text = re.sub(r'\s+', ' ', text)
        return text.upper()
    
    def _clean_document(self, documento):
        """Limpia número de documento y retorna como string numérico."""
        if pd.isna(documento) or documento is None:
            return ""
        doc_clean = re.sub(r'[^\d]', '', str(documento))
        return doc_clean
    
    def _normalize_arm(self, arm_text):
        """Normaliza nombre del brazo usando diccionario."""
        if pd.isna(arm_text) or arm_text is None:
            return "IZQUIERDO"
        
        arm_clean = self._clean_text(arm_text)
        
        if not arm_clean:
            return "IZQUIERDO"
        
        if arm_clean in ARM_DICTIONARY:
            return ARM_DICTIONARY[arm_clean]
        
        for key, value in ARM_DICTIONARY.items():
            if key in arm_clean or arm_clean in key:
                return value
        
        return arm_clean
    
    def _translate_sex(self, sexo):
        """Traduce sexo de MASCULINO/FEMENINO a HOMBRE/MUJER."""
        sexo_upper = self._clean_text(sexo)
        
        if 'MASCULINO' in sexo_upper or sexo_upper == 'M':
            return 'HOMBRE'
        elif 'FEMENINO' in sexo_upper or sexo_upper == 'F':
            return 'MUJER'
        else:
            return sexo_upper
    
    def _format_date(self, date_value):
        """Formatea fecha a DD/MM/YYYY."""
        if pd.isna(date_value) or date_value is None:
            return ""
        
        try:
            if isinstance(date_value, datetime):
                return date_value.strftime('%d/%m/%Y')
            elif isinstance(date_value, str):
                formatted = self._parse_and_format_date(date_value)
                if formatted:
                    return formatted
                return date_value
            return str(date_value)
        except Exception:
            return str(date_value)
    
    def _get_full_name(self, primer_nombre, segundo_nombre):
        """Combina primer y segundo nombre."""
        primer = self._clean_text(primer_nombre)
        segundo = self._clean_text(segundo_nombre)
        
        if segundo:
            return f"{primer} {segundo}"
        return primer
    
    def _parse_biologicos(self, vacuna_texto):
        """Analiza texto de vacuna para extraer biológicos individuales."""
        if pd.isna(vacuna_texto) or vacuna_texto is None:
            return []
        
        vacuna_texto = self._clean_text(vacuna_texto)
        
        if " - " in vacuna_texto:
            parts = vacuna_texto.split(" - ")
        elif "-" in vacuna_texto and not vacuna_texto.startswith("-"):
            parts = vacuna_texto.split("-")
        elif "/" in vacuna_texto:
            parts = vacuna_texto.split("/")
        elif "," in vacuna_texto:
            parts = vacuna_texto.split(",")
        else:
            parts = [vacuna_texto]
        
        biologicos = []
        for part in parts:
            part = part.strip()
            if part:
                biologicos.append(part)
        
        return biologicos
