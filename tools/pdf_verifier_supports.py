import os
import re
import zipfile
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, filedialog
from openpyxl import load_workbook
from typing import List, Dict, Tuple, Set
from ui.toast_notification import (
    ToastManager,
    toast_info,
    toast_success,
    toast_warning,
    toast_error
)


# ==================== UTILIDADES ====================

def extract_document_from_filename(filename: str) -> str:
    """Extrae el documento del nombre de archivo."""
    if not filename:
        return ""
    
    name = os.path.splitext(filename)[0]
    
    if '_' in name:
        last_part = name.split('_')[-1]
    else:
        last_part = name
    
    pattern = r'^(CC|TI|CE|RC|PA|PE|NIT|CD|SC|CN|AS|MS)(\d+)$'
    match = re.match(pattern, last_part.upper())
    
    if match:
        return f"{match.group(1)}{match.group(2)}"
    
    return ""


def extract_document_from_input(text: str) -> str:
    """Extrae el documento del input del usuario."""
    if not text:
        return ""
    
    text_clean = text.strip().upper()
    
    pattern = r'^(CC|TI|CE|RC|PA|PE|NIT|CD|SC|CN|AS|MS)[\s\-_]*(\d+)$'
    match = re.match(pattern, text_clean)
    
    if match:
        return f"{match.group(1)}{match.group(2)}"
    
    return ""


def normalize_document(doc: str) -> str:
    """Normaliza un documento para comparación."""
    if not doc:
        return ""
    return doc.strip().upper()


def classify_files(folder: str) -> Dict[str, List[Tuple[str, str]]]:
    """Clasifica archivos PDF por prefijo CRC/OPF."""
    files = {'CRC': [], 'OPF': [], 'OTHER': []}
    
    if not os.path.exists(folder):
        return files
    
    for f in os.listdir(folder):
        if not f.lower().endswith('.pdf'):
            continue
        
        full_path = os.path.join(folder, f)
        if not os.path.isfile(full_path):
            continue
        
        doc = extract_document_from_filename(f)
        name_upper = f.upper()
        
        if name_upper.startswith('CRC'):
            files['CRC'].append((f, doc))
        elif name_upper.startswith('OPF'):
            files['OPF'].append((f, doc))
        else:
            files['OTHER'].append((f, doc))
    
    return files


def build_document_index(files: List[Tuple[str, str]]) -> Dict[str, str]:
    """Construye índice: documento_normalizado -> nombre_archivo."""
    index = {}
    for filename, doc in files:
        if doc:
            norm_doc = normalize_document(doc)
            if norm_doc and norm_doc not in index:
                index[norm_doc] = filename
    return index


def import_detalle_cargue(filepath: str) -> Tuple[List[str], str]:
    """Importa documentos desde archivo de Detalle de Cargue."""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        header_tipo_id = ws.cell(row=2, column=3).value
        header_id = ws.cell(row=2, column=4).value
        
        if not header_tipo_id or not header_id:
            wb.close()
            return [], "Archivo inválido: No se encontraron los encabezados en la fila 2."
        
        header_tipo_id_clean = str(header_tipo_id).strip().upper()
        header_id_clean = str(header_id).strip().upper()
        
        if "TIPO" not in header_tipo_id_clean or "ID" not in header_tipo_id_clean:
            wb.close()
            return [], f"Encabezado columna C inválido: '{header_tipo_id}'. Se esperaba 'TIPO ID'."
        
        if header_id_clean != "ID":
            wb.close()
            return [], f"Encabezado columna D inválido: '{header_id}'. Se esperaba 'ID'."
        
        docs_set: Set[str] = set()
        
        for row in ws.iter_rows(min_row=3, min_col=3, max_col=4, values_only=True):
            tipo_id, id_num = row
            
            if tipo_id is None and id_num is None:
                continue
            
            if tipo_id and id_num:
                tipo_str = str(tipo_id).strip().upper()
                id_str = str(id_num).strip()
                
                if '.' in id_str:
                    id_str = id_str.split('.')[0]
                
                if tipo_str and id_str:
                    docs_set.add(f"{tipo_str}{id_str}")
        
        wb.close()
        
        if not docs_set:
            return [], "No se encontraron datos válidos en el archivo."
        
        return sorted(docs_set), ""
        
    except Exception as e:
        return [], f"Error al leer el archivo: {e}"


def create_zip_file(folder_path: str, crc_files: List[Tuple[str, str]], opf_files: List[Tuple[str, str]]) -> Tuple[bool, str, int]:
    """
    Crea archivo ZIP con los archivos CRC y OPF.
    
    Retorna: (éxito, mensaje, cantidad_archivos)
    """
    if not crc_files:
        return False, "No hay archivos CRC para comprimir.", 0
    
    folder_name = os.path.basename(folder_path)
    zip_filename = f"{folder_name}.zip"
    zip_path = os.path.join(folder_path, zip_filename)
    
    # Si ya existe, eliminarlo
    if os.path.exists(zip_path):
        try:
            os.remove(zip_path)
        except Exception as e:
            return False, f"No se pudo eliminar ZIP existente: {e}", 0
    
    files_to_add = []
    
    # Recopilar archivos CRC
    for filename, _ in crc_files:
        file_path = os.path.join(folder_path, filename)
        if os.path.exists(file_path) and os.path.isfile(file_path):
            files_to_add.append((file_path, filename))
    
    # Recopilar archivos OPF
    for filename, _ in opf_files:
        file_path = os.path.join(folder_path, filename)
        if os.path.exists(file_path) and os.path.isfile(file_path):
            files_to_add.append((file_path, filename))
    
    if not files_to_add:
        return False, "No se encontraron archivos válidos para comprimir.", 0
    
    try:
        with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
            for file_path, arcname in files_to_add:
                zf.write(file_path, arcname=arcname)
        
        # Verificar que el ZIP se creó correctamente
        if not os.path.exists(zip_path):
            return False, "El archivo ZIP no se creó.", 0
        
        # Verificar que el ZIP es válido
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                if zf.testzip() is not None:
                    return False, "El archivo ZIP está corrupto.", 0
        except zipfile.BadZipFile:
            return False, "El archivo ZIP está corrupto.", 0
        
        return True, zip_filename, len(files_to_add)
        
    except Exception as e:
        # Limpiar archivo corrupto si existe
        if os.path.exists(zip_path):
            try:
                os.remove(zip_path)
            except:
                pass
        return False, f"Error creando ZIP: {e}", 0


# ==================== LÓGICA DE VERIFICACIÓN ====================

def verify_documents(input_docs: List[str], crc_files: List[Tuple[str, str]], opf_files: List[Tuple[str, str]]) -> Tuple[List[Dict], bool]:
    """
    Verifica documentos. REQUIERE MÍNIMO 2 DE LAS 3 FUENTES.
    
    Retorna: (resultados, puede_crear_zip)
    """
    results = []
    
    crc_index = build_document_index(crc_files)
    opf_index = build_document_index(opf_files)
    
    input_set = {}
    crc_set = set(crc_index.keys())
    opf_set = set(opf_index.keys())
    
    invalid_inputs = []
    for doc in input_docs:
        doc_extracted = extract_document_from_input(doc)
        if doc_extracted:
            norm_doc = normalize_document(doc_extracted)
            if norm_doc:
                input_set[norm_doc] = doc_extracted
        else:
            invalid_inputs.append(doc.strip())
    
    has_input = len(input_set) > 0 or len(invalid_inputs) > 0
    has_crc = len(crc_files) > 0
    has_opf = len(opf_files) > 0
    
    sources = sum([has_input, has_crc, has_opf])
    
    if sources < 2:
        return [], False
    
    can_zip = has_crc
    
    for invalid_doc in invalid_inputs:
        results.append({
            'documento': invalid_doc,
            'crc_archivo': '',
            'opf_archivo': '',
            'estado': "❌ Formato inválido"
        })
    
    # Caso 1: INPUT + CRC (sin OPF)
    if has_input and has_crc and not has_opf:
        all_docs_norm = sorted(set(input_set.keys()) | crc_set)
        
        for norm_doc in all_docs_norm:
            orig_doc = input_set.get(norm_doc, '')
            in_input = norm_doc in input_set
            crc_file = crc_index.get(norm_doc, '')
            
            if in_input and crc_file:
                pass
            elif in_input and not crc_file:
                results.append({
                    'documento': orig_doc,
                    'crc_archivo': '',
                    'opf_archivo': '',
                    'estado': "❌ Sin soporte CRC"
                })
            elif not in_input and crc_file:
                results.append({
                    'documento': '',
                    'crc_archivo': crc_file,
                    'opf_archivo': '',
                    'estado': "⚠️ No esta en Listado"
                })
    
    # Caso 2: INPUT + OPF (sin CRC)
    elif has_input and has_opf and not has_crc:
        can_zip = False
        all_docs_norm = sorted(set(input_set.keys()) | opf_set)
        
        for norm_doc in all_docs_norm:
            orig_doc = input_set.get(norm_doc, '')
            in_input = norm_doc in input_set
            opf_file = opf_index.get(norm_doc, '')
            
            if in_input and opf_file:
                pass
            elif in_input and not opf_file:
                results.append({
                    'documento': orig_doc,
                    'crc_archivo': '',
                    'opf_archivo': '',
                    'estado': "❌ Sin soporte OPF"
                })
            elif not in_input and opf_file:
                results.append({
                    'documento': '',
                    'crc_archivo': '',
                    'opf_archivo': opf_file,
                    'estado': "⚠️ No esta en Listado"
                })
    
    # Caso 3: CRC + OPF (sin INPUT)
    elif not has_input and has_crc and has_opf:
        all_docs = sorted(crc_set | opf_set)
        for norm_doc in all_docs:
            crc_file = crc_index.get(norm_doc, '')
            opf_file = opf_index.get(norm_doc, '')
            
            if crc_file and opf_file:
                pass
            elif crc_file and not opf_file:
                results.append({
                    'documento': 'SOLO ARCHIVOS',
                    'crc_archivo': crc_file,
                    'opf_archivo': '',
                    'estado': "⚠️ Falta OPF"
                })
            elif opf_file and not crc_file:
                results.append({
                    'documento': 'SOLO ARCHIVOS',
                    'crc_archivo': '',
                    'opf_archivo': opf_file,
                    'estado': "⚠️ Falta CRC"
                })
    
    # Caso 4: INPUT + CRC + OPF
    elif has_input and has_crc and has_opf:
        all_docs_norm = sorted(set(input_set.keys()) | crc_set | opf_set)
        
        for norm_doc in all_docs_norm:
            orig_doc = input_set.get(norm_doc, '')
            in_input = norm_doc in input_set
            crc_file = crc_index.get(norm_doc, '')
            opf_file = opf_index.get(norm_doc, '')
            
            if in_input and crc_file and opf_file:
                pass
            elif in_input and crc_file and not opf_file:
                results.append({
                    'documento': orig_doc,
                    'crc_archivo': crc_file,
                    'opf_archivo': '',
                    'estado': "⚠️ Falta OPF"
                })
            elif in_input and opf_file and not crc_file:
                results.append({
                    'documento': orig_doc,
                    'crc_archivo': '',
                    'opf_archivo': opf_file,
                    'estado': "⚠️ Falta CRC"
                })
            elif in_input and not crc_file and not opf_file:
                results.append({
                    'documento': orig_doc,
                    'crc_archivo': '',
                    'opf_archivo': '',
                    'estado': "❌ Sin soportes CRC/OPF"
                })
            elif not in_input and crc_file and opf_file:
                results.append({
                    'documento': '',
                    'crc_archivo': crc_file,
                    'opf_archivo': opf_file,
                    'estado': "⚠️ No estan en Listado"
                })
            elif not in_input and crc_file:
                results.append({
                    'documento': '',
                    'crc_archivo': crc_file,
                    'opf_archivo': '',
                    'estado': "⚠️ Solo CRC"
                })
            elif not in_input and opf_file:
                results.append({
                    'documento': '',
                    'crc_archivo': '',
                    'opf_archivo': opf_file,
                    'estado': "⚠️ Solo OPF"
                })
    
    return results, can_zip


# ==================== APLICACIÓN GUI ====================

class PDFVerifierSupportsApp(ctk.CTkFrame):
    """Aplicación para verificar PDFs CRC/OPF."""
    
    def __init__(self, master, go_home=None):
        super().__init__(master)
        self.folder_path = ""
        self.last_files = {'CRC': [], 'OPF': []}
        self.can_create_zip = False
        self._build_ui()
    
    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        
        self._build_top_panel()
        self._build_input_panel()
        self._build_results_panel()
        self._build_log_panel()
    
    def _build_top_panel(self):
        panel = ctk.CTkFrame(self)
        panel.grid(row=0, column=0, sticky="ew", padx=6, pady=6)
        panel.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(panel, text="📁 Carpeta:", width=80, anchor="w", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=(6,4), pady=4, sticky="w")
        self.folder_entry = ctk.CTkEntry(panel, placeholder_text="Selecciona carpeta con PDFs (CRC/OPF)")
        self.folder_entry.grid(row=0, column=1, padx=4, pady=4, sticky="ew")
        ctk.CTkButton(panel, text="Seleccionar", width=120, command=self._select_folder).grid(row=0, column=2, padx=(4,6), pady=4)
        
        btn_frame = ctk.CTkFrame(panel, fg_color="transparent")
        btn_frame.grid(row=1, column=0, columnspan=3, sticky="ew", padx=6, pady=(4,6))
        btn_frame.grid_columnconfigure((0,1,2), weight=1)
        
        self.verify_btn = ctk.CTkButton(btn_frame, text="🔍 Verificar", command=self._verify, fg_color="#1f6feb")
        self.verify_btn.grid(row=0, column=0, padx=4, pady=4, sticky="ew")
        
        self.zip_btn = ctk.CTkButton(btn_frame, text="📦 Convertir a ZIP", command=self._create_zip, fg_color="#27ae60", state="disabled")
        self.zip_btn.grid(row=0, column=1, padx=4, pady=4, sticky="ew")
        
        self.clear_btn = ctk.CTkButton(btn_frame, text="🧹 Limpiar", command=self._clear, fg_color="#e74c3c")
        self.clear_btn.grid(row=0, column=2, padx=4, pady=4, sticky="ew")
    
    def _build_input_panel(self):
        panel = ctk.CTkFrame(self)
        panel.grid(row=1, column=0, sticky="ew", padx=6, pady=(0,6))
        panel.grid_columnconfigure(0, weight=1)
        
        header = ctk.CTkFrame(panel, fg_color="transparent")
        header.pack(fill="x", padx=6, pady=(6,4))
        
        ctk.CTkLabel(header, text="📝 Documentos (opcional) - Ej: CC2579081", font=ctk.CTkFont(size=12, weight="bold")).pack(side="left")
        ctk.CTkButton(header, text="📥 Importar Detalle Cargue", width=180, command=self._import_detalle_cargue).pack(side="right")
        
        self.input_text = ctk.CTkTextbox(panel, height=80, font=ctk.CTkFont(family="Consolas", size=11))
        self.input_text.pack(fill="x", padx=6, pady=(0,6))
    
    def _build_results_panel(self):
        panel = ctk.CTkFrame(self)
        panel.grid(row=2, column=0, sticky="nsew", padx=6, pady=(0,6))
        panel.grid_columnconfigure(0, weight=1)
        panel.grid_rowconfigure(1, weight=1)
        
        header = ctk.CTkFrame(panel, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=6, pady=(6,4))
        header.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(header, text="📊 Incompletos/Faltantes", font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=0, sticky="w")
        self.count_lbl = ctk.CTkLabel(header, text="0 registros", text_color="gray", anchor="e")
        self.count_lbl.grid(row=0, column=1, sticky="e")
        
        table_frame = ctk.CTkFrame(panel)
        table_frame.grid(row=1, column=0, sticky="nsew", padx=6, pady=(0,6))
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)
        
        self._setup_table_style()
        
        cols = ("documento", "crc", "opf", "estado")
        self.table = ttk.Treeview(table_frame, columns=cols, show="headings", style="Dark.Treeview")
        
        self.table.heading("documento", text="Documento")
        self.table.heading("crc", text="Archivo CRC")
        self.table.heading("opf", text="Archivo OPF")
        self.table.heading("estado", text="Estado")
        
        self.table.column("documento", width=150)
        self.table.column("crc", width=280)
        self.table.column("opf", width=280)
        self.table.column("estado", width=180)
        
        scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.table.yview)
        scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.table.xview)
        self.table.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.table.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
    
    def _build_log_panel(self):
        panel = ctk.CTkFrame(self)
        panel.grid(row=3, column=0, sticky="ew", padx=6, pady=(0,6))
        panel.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(panel, text="📋 Registro", font=ctk.CTkFont(weight="bold"), anchor="w").grid(row=0, column=0, sticky="w", padx=6, pady=(6,2))
        
        self.log_text = ctk.CTkTextbox(panel, height=70, font=ctk.CTkFont(family="Consolas", size=11))
        self.log_text.grid(row=1, column=0, sticky="ew", padx=6, pady=(0,6))
        self.log_text.configure(state="disabled")
        
        self._log("INFO", "Archivos: CRC_xxx_xxx_CC123.pdf / OPF_xxx_xxx_CC123.pdf")
    
    def _setup_table_style(self):
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Dark.Treeview", background="#2b2b2b", foreground="white", fieldbackground="#2b2b2b", borderwidth=0, rowheight=26)
        style.configure("Dark.Treeview.Heading", background="#3b3b3b", foreground="white", borderwidth=1)
        style.map("Dark.Treeview", background=[("selected", "#1f6feb")], foreground=[("selected", "white")])
    #_log1
    def _log(self, level: str, message: str):
        colors = {"INFO": "#7f8c8d", "SUCCESS": "#27ae60", "WARNING": "#f39c12", "ERROR": "#e74c3c"}
        
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{level}] {message}\n")
        
        last_line = self.log_text.index("end-2l")
        self.log_text.tag_add(level, last_line, "end-1c")
        self.log_text.tag_config(level, foreground=colors.get(level, "white"))
        
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.update()
    
    def _select_folder(self):
        folder = filedialog.askdirectory(title="Seleccionar carpeta")
        if folder:
            self.folder_path = folder
            self.folder_entry.delete(0, "end")
            self.folder_entry.insert(0, folder)
            self._log("INFO", f"Carpeta: {folder}")
            self._disable_zip()
    
    def _import_detalle_cargue(self):
        """Importa documentos desde archivo de Detalle de Cargue."""
        path = filedialog.askopenfilename(
            title="Seleccionar Detalle de Cargue",
            filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        
        if not path:
            return
        
        toast_info(f"Importando: {os.path.basename(path)}", 3)
        self.update()
        
        docs, error = import_detalle_cargue(path)
        
        if error:

            self.after(200, lambda: toast_error(error))
            self.after(700, lambda: toast_warning("Por favor importe un archivo de Detalle de Cargue válido.", 5))
            return
        
        self.input_text.delete("1.0", "end")
        self.input_text.insert("1.0", "\n".join(docs))
        self.after(200, lambda: toast_success(f"Importado correctamente.\n{len(docs)} documentos cargados."))
        self._disable_zip()
    
    def _disable_zip(self):
        """Desactiva el botón ZIP."""
        self.can_create_zip = False
        self.zip_btn.configure(state="disabled")
    
    def _enable_zip(self):
        """Activa el botón ZIP."""
        self.can_create_zip = True
        self.zip_btn.configure(state="normal")
    
    def _verify(self):
        if not self.folder_path:
            toast_warning("Selecciona una carpeta.", 3)
            return
        
        for item in self.table.get_children():
            self.table.delete(item)
        
        self._disable_zip()
        
        input_text = self.input_text.get("1.0", "end").strip()
        input_docs = [line.strip() for line in input_text.splitlines() if line.strip()]
    
        toast_info("Clasificando archivos...", 3)
        files = classify_files(self.folder_path)
        
        self.last_files = files
        
        has_input = len(input_docs) > 0
        has_crc = len(files['CRC']) > 0
        has_opf = len(files['OPF']) > 0
        sources = sum([has_input, has_crc, has_opf])
        
        toast_info(f"DATOS\nDOC: {len(input_docs)}\nCRC: {len(files['CRC'])}\nOPF: {len(files['OPF'])}", 10)
        
        if sources < 2:
            toast_error("❌ Ingrese una carpeta con los archivos o inserte documentos en el panel.", 5)
            self.count_lbl.configure(text="0 registros")
            return
        
        toast_info("Verificando...", 3)
        results, can_zip = verify_documents(input_docs, files['CRC'], files['OPF'])
        
        for r in results:
            self.table.insert("", "end", values=(
                r['documento'],
                r['crc_archivo'],
                r['opf_archivo'],
                r['estado']
            ))
        
        if len(results) == 0:
            toast_success("✅ Todos los documentos son compatibles.", 5)
            if can_zip:
                self._enable_zip()
            else:
                toast_warning("⚠️ No se puede crear ZIP: Solo hay OPF (requiere CRC).", 5)
        else:
            toast_warning(f"⚠️ {len(results)} incompatibles.", 0)
        
        self.count_lbl.configure(text=f"{len(results)} incompletos")
    
    def _create_zip(self):
        """Crea archivo ZIP con todos los archivos CRC y OPF."""
        if not self.can_create_zip:
            self._log("WARNING", "Primero verifica que todo esté completo.")
            return
        
        if not self.folder_path:
            toast_warning("Debes seleccionar una carpeta.", 3)
            return
        
        crc_files = self.last_files.get('CRC', [])
        opf_files = self.last_files.get('OPF', [])
        
        toast_info("Creando ZIP...", 3)
        self.update()
        
        success, message, count = create_zip_file(self.folder_path, crc_files, opf_files)
        
        if success:
            toast_success(f"✅ ZIP creado: {message}", 5)
            self._log("SUCCESS", f"📦 {count} archivos ({len(crc_files)} CRC + {len(opf_files)} OPF)")
        else:
            self._log("ERROR", message)
    
    def _clear(self):
        self.input_text.delete("1.0", "end")
        for item in self.table.get_children():
            self.table.delete(item)
        self.count_lbl.configure(text="0 registros")
        self._disable_zip()
        toast_info("Limpiado.", 3)


if __name__ == "__main__":
    root = ctk.CTk()
    root.title("Verificador CRC/OPF")
    root.geometry("1000x700")
    root.minsize(900, 600)
    app = PDFVerifierSupportsApp(root)
    app.pack(fill="both", expand=True)
    root.mainloop()